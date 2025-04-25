from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login,logout
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.contrib.admin.views.decorators import staff_member_required
from users.models import CustomUser
from django.core.paginator import Paginator
from users.models import CustomUser
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.db.models import Sum
from datetime import timedelta
from django.db.models.functions import ExtractHour, ExtractMonth,TruncDate
from orders.models import Order, OrderItem
from django.utils import timezone
from django.db.models import F


def BariqadminLogin(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            if user.is_staff or user.is_superuser:
                # Log the user in
                login(request, user)
                messages.success(request, 'You have successfully logged in as an admin!')
                return redirect('admin_dashboard')  
            else:
                
                messages.error(request, 'You do not have permission to access the admin panel.')
                return redirect('adminlogin') 
        else:
    
            messages.error(request, 'Invalid username or password.')
            return redirect('adminlogin')  

   
    return render(request, 'adminlogin.html')
@staff_member_required
def admin_logout(request):
    logout(request)
    return redirect('admin_login')



#_______________________________________________________________________________________________________________________________#



@staff_member_required
def Admin_Dashboard(request):
    # Get time filter and date range from request
    filter_type = request.GET.get('filter', 'monthly')
    start_date_param = request.GET.get('start_date')
    end_date_param = request.GET.get('end_date')
    
    today = timezone.now()  # Use timezone-aware datetime
    error_message = None
    is_custom_range = False
    
    # Handle custom date range
    if start_date_param and end_date_param:
        try:
            start_date = timezone.make_aware(datetime.strptime(start_date_param, '%Y-%m-%d'))
            end_date = timezone.make_aware(datetime.strptime(end_date_param, '%Y-%m-%d'))
            # Add time to include the entire end day
            end_date = end_date.replace(hour=23, minute=59, second=59)
            # Validate date range
            if end_date < start_date:
                error_message = "End date cannot be before start date."
                start_date, end_date = get_default_dates(filter_type, today)
            elif end_date > today:
                error_message = "End date cannot be in the future."
                end_date = today
            else:
                is_custom_range = True
        except ValueError:
            error_message = "Invalid date format. Using default dates."
            start_date, end_date = get_default_dates(filter_type, today)
    else:
        start_date, end_date = get_default_dates(filter_type, today)
    
    if error_message:
        messages.error(request, error_message)
    
    # Filtered Orders with proper date range
    orders = Order.objects.filter(order_date__gte=start_date, order_date__lte=end_date)

    # Summary Stats (calculated from filtered orders)
    total_orders = orders.count()
    total_sales = orders.aggregate(total=Sum("total_amount"))["total"] or 0
    total_customers = orders.values("user").distinct().count()

    # Sales Chart Data with proper grouping based on date range
    date_diff = (end_date - start_date).days
    if date_diff <= 1:
        # Group by hour for single-day range
        sales_data = (
            orders.annotate(
                hour=ExtractHour('order_date')
            ).values('hour')
            .annotate(
                order_date__date=F('hour'),
                total_sales=Sum("total_amount")
            ).order_by("hour")
        )
    elif date_diff > 365:
        # Group by month for long ranges
        sales_data = (
            orders.annotate(
                month=ExtractMonth('order_date')
            ).values('month')
            .annotate(
                order_date__date=F('month'),
                total_sales=Sum("total_amount")
            ).order_by("month")
        )
    else:
        # Group by date for other ranges
        sales_data = (
            orders.annotate(
                date=TruncDate('order_date')
            ).values('date')
            .annotate(
                order_date__date=F('date'),
                total_sales=Sum("total_amount")
            ).order_by("date")
        )

    # Best Selling Products
    best_products = (
        OrderItem.objects.filter(order__order_date__gte=start_date, order__order_date__lte=end_date)
        .values("product_variant__product__name")
        .annotate(total_quantity=Sum("quantity"))
        .order_by("-total_quantity")[:5]
    )

    # Best Selling Categories
    best_categories = (
        OrderItem.objects.filter(order__order_date__gte=start_date, order__order_date__lte=end_date)
        .values("product_variant__product__category__name")
        .annotate(total_quantity=Sum("quantity"))
        .order_by("-total_quantity")[:5]
    )

    context = {
        "sales_data": sales_data,
        "best_products": best_products,
        "best_categories": best_categories,
        "total_orders": total_orders,
        "total_sales": total_sales,
        "total_customers": total_customers,
        "filter_type": filter_type,
        "start_date": start_date,
        "end_date": end_date,
        "is_custom_range": is_custom_range,
    }
    return render(request, "admin-dashboard.html", context)

def get_default_dates(filter_type, today):
    """Helper function to get default date ranges based on filter type"""
    if filter_type == 'daily':
        start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif filter_type == 'weekly':
        start_date = today - timedelta(days=today.weekday())
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = today
    elif filter_type == 'yearly':
        start_date = today.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end_date = today
    else:  # Monthly (default)
        start_date = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_date = today
    
    return start_date,end_date     




@staff_member_required
def user_management(request):
    users_list = CustomUser.objects.filter(is_superuser=False).order_by('-date_of_joining')
    paginator = Paginator(users_list, 10)
    page = request.GET.get('page')
    users = paginator.get_page(page)
    context = {
        'users': users,
        'total_users': users_list.count(),
    }
    return render(request, 'user_management.html', context)

@csrf_exempt
@require_POST
def toggle_user_status(request):
    user_id = request.POST.get('user_id')
    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'User ID is required'}, status=400)

    try:
        user = CustomUser.objects.get(id=user_id)
        user.is_blocked = not user.is_blocked
        user.save()
        return JsonResponse({'status': 'success', 'is_blocked': user.is_blocked})
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'User not found'}, status=404)

@staff_member_required
def offers(request):
    return render(request,'offers.html')








#_________________________________________________________________________________________________________________________#
from datetime import datetime, timedelta
from django.utils.timezone import now
from django.http import HttpResponse
from decimal import Decimal
from django.db.models import Sum, Count, Q, F


# PDF report imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, Table, TableStyle, SimpleDocTemplate, Spacer

# Excel report imports
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter



@staff_member_required
def sales_report(request):
    filter_type    = request.GET.get('filter', 'daily')
    start_date_str = request.GET.get('start_date')
    end_date_str   = request.GET.get('end_date')
    today          = now().date()

    # 1. Determine start_date / end_date exactly as before
    if filter_type == 'daily':
        start_date = end_date = today
    elif filter_type == 'weekly':
        start_date = today - timedelta(days=today.weekday())
        end_date   = today
    elif filter_type == 'monthly':
        start_date = today.replace(day=1)
        end_date   = today
    elif filter_type == 'yearly':
        start_date = today.replace(month=1, day=1)
        end_date   = today
    elif filter_type == 'custom' and start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date   = datetime.strptime(end_date_str,   '%Y-%m-%d').date()
            if end_date < start_date:
                start_date = end_date = today
        except ValueError:
            start_date = end_date = today
    else:
        start_date = end_date = today

    # 2. Base queryset: all orders in date range
    base_orders = Order.objects.filter(order_date__date__range=(start_date, end_date))

    # 3. Annotate how many items in each, and how many of those are cancelled/refunded
    orders = base_orders.annotate(
        total_items     = Count('items'),
        cancelled_items = Count('items', filter=Q(items__status__in=['cancelled','refunded'])),
    ).filter(
        # keep any order that has at least one non-cancelled item
        ~Q(cancelled_items=F('total_items'))
    )

    # 4. Calculate sums *only* over the non-cancelled/refunded items
    non_cancel_q = Q(order__in=orders) & ~Q(status__in=['cancelled','refunded'])
    item_aggregates = OrderItem.objects.filter(non_cancel_q).aggregate(
        subtotal             = Sum('price'),
        product_discount     = Sum('discount'),
    )

    # 5. Coupon and shipping are per‐order, include only orders we kept
    order_aggregates = orders.aggregate(
        coupon_discount = Sum('coupon_discount'),
        shipping_cost   = Sum('shipping_cost'),
    )

    # 6. Totals
    total_sales_count     = orders.count()
    total_subtotal        = item_aggregates['subtotal'] or 0
    total_product_discount= item_aggregates['product_discount'] or 0
    total_coupon_discount = order_aggregates['coupon_discount'] or 0
    total_shipping        = order_aggregates['shipping_cost'] or 0

    # 7. Net sales
    net_sales = (
        total_subtotal
        - total_product_discount
        - total_coupon_discount
        + total_shipping
    )

    context = {
        'orders': orders,
        'total_sales_count': total_sales_count,
        'total_sales_amount': total_subtotal,
        'total_product_discount': total_product_discount,
        'total_coupon_discount': total_coupon_discount,
        'total_shipping': total_shipping,
        'net_sales': net_sales,
        'start_date': start_date,
        'end_date': end_date,
        'filter_type': filter_type,
    }
    return render(request, 'sales_report.html', context)



@staff_member_required
def download_pdf(request):
    # Get date parameters from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        start_date = end_date = now().date()

    # Query orders based on date range
    orders = Order.objects.filter(
        order_date__date__range=[start_date, end_date]
    ).exclude(
        items__status__in=['cancelled', 'refunded']
    ).distinct()

    # Set up HTTP response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Bariq_Fashions_Sales_Report_{start_date}_to_{end_date}.pdf"'

    # Define colors for better design
    BRAND_COLOR = colors.HexColor('#3498DB')  # Bright blue
    HEADER_COLOR = colors.HexColor('#2E86C1')  # Slightly darker blue
    ALTERNATE_ROW_COLOR = colors.HexColor('#F2F4F4')  # Light gray for zebra striping
    TOTAL_ROW_COLOR = colors.HexColor('#D4E6F1')  # Light blue for totals row

    # A4 dimensions
    width, height = A4

    # Define margins and dimensions
    LEFT_MARGIN = 20 * mm
    RIGHT_MARGIN = width - 20 * mm
    TABLE_WIDTH = width - (LEFT_MARGIN * 2)
    ROW_HEIGHT = 10 * mm  # Increased for better readability

    # Create a SimpleDocTemplate for better multi-page handling
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=LEFT_MARGIN,
        rightMargin=LEFT_MARGIN,
        topMargin=20 * mm,
        bottomMargin=20 * mm
    )
    
    # Create elements list for the document
    elements = []
    
    # Add title and header as a table for consistent positioning
    header_data = [
        [Paragraph("<font size='18' color='white'><b>Bariq Fashions</b></font>", 
                 ParagraphStyle(name='Center', alignment=1))],
        [Paragraph(f"<font size='14' color='white'><b>Sales Report</b></font>", 
                 ParagraphStyle(name='Center', alignment=1))],
        [Paragraph(f"<font size='12' color='white'>{start_date} to {end_date}</font>", 
                 ParagraphStyle(name='Center', alignment=1))]
    ]
    
    header_table = Table(header_data, colWidths=[TABLE_WIDTH])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), BRAND_COLOR),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
    ]))
    
    elements.append(header_table)
    elements.append(Spacer(1, 15 * mm))
    
    # Properly aligned column widths
    col_widths = [
        30 * mm,   # Order ID
        25 * mm,   # Date
        30 * mm,   # Subtotal
        30 * mm,   # Product Discount
        30 * mm,   # Coupon Discount
        22 * mm,   # Shipping
        23 * mm    # Total
    ]
    
    # Use Paragraph objects for headers to ensure proper alignment and wrapping
    header_style = ParagraphStyle(
        'HeaderStyle',
        fontName='Helvetica-Bold',
        fontSize=10,
        alignment=1,  # Center alignment
        textColor=colors.white
    )
    
    # Setup data for the report - using Paragraph objects for headers
    headers = [
        Paragraph("Order ID", header_style),
        Paragraph("Date", header_style),
        Paragraph("Subtotal (INR)", header_style),
        Paragraph("Product Discount", header_style),
        Paragraph("Coupon Discount", header_style),
        Paragraph("Shipping", header_style),
        Paragraph("Total (INR)", header_style)
    ]
    
    data = [headers]
    
    # Track totals
    total_subtotal = Decimal('0.00')
    total_product_discount = Decimal('0.00')
    total_coupon_discount = Decimal('0.00')
    total_shipping = Decimal('0.00')
    total_amount = Decimal('0.00')
    
    # Number format function for consistency
    def format_number(value):
        return f"{value:,.2f}" if value else "0.00"
    
    # Regular cell style for right-aligned numbers
    num_style = ParagraphStyle(
        'NumberStyle', 
        fontName='Helvetica',
        fontSize=10,
        alignment=2  # Right alignment
    )
    
    # Text cell style for left-aligned text
    text_style = ParagraphStyle(
        'TextStyle',
        fontName='Helvetica',
        fontSize=10,
        alignment=0  # Left alignment
    )
    
    # Collect order data
    for order in orders:
        subtotal = order.subtotal
        product_discount = order.product_discount
        coupon_discount = order.coupon_discount
        shipping = order.shipping_cost
        total = order.total_amount
        
        row = [
            Paragraph(str(order.order_id), text_style),
            Paragraph(order.order_date.strftime('%Y-%m-%d'), text_style),
            Paragraph(format_number(subtotal), num_style),
            Paragraph(format_number(product_discount), num_style),
            Paragraph(format_number(coupon_discount), num_style),
            Paragraph(format_number(shipping), num_style),
            Paragraph(format_number(total), num_style)
        ]
        data.append(row)
        
        # Update totals
        total_subtotal += subtotal
        total_product_discount += product_discount
        total_coupon_discount += coupon_discount
        total_shipping += shipping
        total_amount += total
    
    # Total row style
    total_style = ParagraphStyle(
        'TotalStyle',
        fontName='Helvetica-Bold',
        fontSize=10,
        alignment=2  # Right alignment
    )
    
    total_label_style = ParagraphStyle(
        'TotalLabelStyle',
        fontName='Helvetica-Bold',
        fontSize=10,
        alignment=0  # Left alignment
    )
    
    # Add totals row
    totals_row = [
        Paragraph("TOTAL", total_label_style),
        Paragraph("", total_label_style),
        Paragraph(format_number(total_subtotal), total_style),
        Paragraph(format_number(total_product_discount), total_style),
        Paragraph(format_number(total_coupon_discount), total_style),
        Paragraph(format_number(total_shipping), total_style),
        Paragraph(format_number(total_amount), total_style)
    ]
    data.append(totals_row)
    
    # Create and style the main table
    main_table = Table(data, colWidths=col_widths, repeatRows=1)
    
    # Apply styling to the table
    style = TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), HEADER_COLOR),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, ALTERNATE_ROW_COLOR]),
        
        # Total row styling
        ('BACKGROUND', (0, -1), (-1, -1), TOTAL_ROW_COLOR),
        
        # Grid styling - lighter grid lines for better appearance
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        
        # Spacing
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ])
    
    # Add extra styling for total row
    if len(data) > 2:  # If we have data rows plus header
        style.add('LINEBELOW', (0, -2), (-1, -2), 1, colors.black)  # Line above totals
    
    main_table.setStyle(style)
    elements.append(main_table)
    
    # Add footer with page numbers
    def add_page_number(canvas, doc):
        canvas.saveState()
        page_num = canvas.getPageNumber()
        text = f"Page {page_num}"
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(RIGHT_MARGIN, 15 * mm, text)
        canvas.drawString(LEFT_MARGIN, 15 * mm, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        canvas.restoreState()
    
    # Build the document with all elements
    doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
    return response



@staff_member_required
def download_excel(request):
    # Get date parameters from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        start_date = end_date = now().date()

    # Query orders based on date range
    orders = Order.objects.filter(
        order_date__date__range=[start_date, end_date]
    ).exclude(
        items__status__in=['cancelled', 'refunded']
    ).distinct()

    # Create workbook and set active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Define styles
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    subtitle_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    cell_font = Font(name='Calibri', size=11)
    total_font = Font(name='Calibri', size=12, bold=True)
    
    title_fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
    header_fill = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')
    alternate_fill = PatternFill(start_color='F2F4F4', end_color='F2F4F4', fill_type='solid')
    total_fill = PatternFill(start_color='D4E6F1', end_color='D4E6F1', fill_type='solid')
    
    border = Border(
        left=Side(style='thin', color='CCCCCC'), 
        right=Side(style='thin', color='CCCCCC'), 
        top=Side(style='thin', color='CCCCCC'), 
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    
    # Define named styles for reuse
    currency_style = NamedStyle(name='currency_style')
    currency_style.font = cell_font
    currency_style.border = border
    currency_style.alignment = right_alignment
    currency_style.number_format = '#,##0.00'
    wb.add_named_style(currency_style)
    
    date_style = NamedStyle(name='date_style')
    date_style.font = cell_font
    date_style.border = border
    date_style.alignment = center_alignment
    date_style.number_format = 'YYYY-MM-DD'
    wb.add_named_style(date_style)
    
    text_style = NamedStyle(name='text_style')
    text_style.font = cell_font
    text_style.border = border
    text_style.alignment = left_alignment
    wb.add_named_style(text_style)
    
    # Create title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = "Bariq Fashions"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_alignment
    ws.row_dimensions[1].height = 30
    
    # Create subtitle
    ws.merge_cells('A2:G2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Sales Report ({start_date} to {end_date})"
    subtitle_cell.font = subtitle_font
    subtitle_cell.fill = title_fill
    subtitle_cell.alignment = center_alignment
    ws.row_dimensions[2].height = 25
    
    # Add empty row for spacing
    ws.row_dimensions[3].height = 10
    
    # Add headers
    headers = ["Order ID", "Date", "Subtotal (INR)", "Product Discount", 
               "Coupon Discount", "Shipping", "Total (INR)"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    ws.row_dimensions[4].height = 20
    
    # Track totals
    total_subtotal = Decimal('0.00')
    total_product_discount = Decimal('0.00')
    total_coupon_discount = Decimal('0.00')
    total_shipping = Decimal('0.00')
    total_amount = Decimal('0.00')
    
    # Add data rows
    row_num = 5
    for i, order in enumerate(orders):
        # Get order values
        subtotal = order.subtotal
        product_discount = order.product_discount
        coupon_discount = order.coupon_discount
        shipping = order.shipping_cost
        total = order.total_amount
        
        # Apply alternating row colors
        row_fill = alternate_fill if i % 2 == 1 else None
        
        # Add row data
        ws.cell(row=row_num, column=1, value=str(order.order_id)).style = text_style
        ws.cell(row=row_num, column=2, value=order.order_date.date()).style = date_style
        ws.cell(row=row_num, column=3, value=float(subtotal)).style = currency_style
        ws.cell(row=row_num, column=4, value=float(product_discount)).style = currency_style
        ws.cell(row=row_num, column=5, value=float(coupon_discount)).style = currency_style
        ws.cell(row=row_num, column=6, value=float(shipping)).style = currency_style
        ws.cell(row=row_num, column=7, value=float(total)).style = currency_style
        
        # Apply row fill if needed
        if row_fill:
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).fill = row_fill
        
        # Update totals
        total_subtotal += subtotal
        total_product_discount += product_discount
        total_coupon_discount += coupon_discount
        total_shipping += shipping
        total_amount += total
        
        row_num += 1
    
    # Add total row
    total_row = row_num
    
    # Add a thicker border above the total row
    thick_top_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Add total label
    total_label = ws.cell(row=total_row, column=1, value="TOTAL")
    total_label.font = total_font
    total_label.fill = total_fill
    total_label.border = thick_top_border
    total_label.alignment = left_alignment
    
    # Add empty date cell
    empty_cell = ws.cell(row=total_row, column=2, value="")
    empty_cell.font = total_font
    empty_cell.fill = total_fill
    empty_cell.border = thick_top_border
    
    # Add total values - use actual calculated totals rather than formulas
    # This ensures the values are correct even if Excel recalculates
    total_cells = [
        (3, float(total_subtotal)),
        (4, float(total_product_discount)),
        (5, float(total_coupon_discount)),
        (6, float(total_shipping)),
        (7, float(total_amount))
    ]
    
    for col, value in total_cells:
        cell = ws.cell(row=total_row, column=col, value=value)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thick_top_border
        cell.alignment = right_alignment
        cell.number_format = '#,##0.00'
    
    # Set column widths
    column_widths = [20, 15, 18, 18, 18, 15, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze the header row
    ws.freeze_panes = 'A5'
    
    # Add gridlines and adjust print settings
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # Add footer
    footer_row = total_row + 2
    ws.cell(row=footer_row, column=1, value=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=footer_row, column=1).font = Font(name='Calibri', size=9, color='777777')
    
    # Create HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Bariq_Fashions_Sales_Report_{start_date}_to_{end_date}.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    return response